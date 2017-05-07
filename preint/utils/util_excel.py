# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
import xlrd, xlwt

class XlsxHelper():
    def __init__(self,filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active

    def activeSheet(self, sheetname):
        self.ws = self.wb[sheetName]

    def write(self, cell, data):
        # Data can be directly to cells
        self.ws[cell] = data
        # Rows can also be appended
        # Save the file
    def getCell(self, cellname):
        return self.ws[cellname]
    
    def save(self, filename):
        self.wb.save(filename)


class XlsHelper():
    def read_excel():

        workbook = xlrd.open_workbook(r'F:\demo.xlsx')

        print workbook.sheet_names() # [u'sheet1', u'sheet2']
        sheet2_name = workbook.sheet_names()[1]
  
        sheet2 = workbook.sheet_by_index(1) # sheet索引从0开始
        sheet2 = workbook.sheet_by_name('sheet2')

        print sheet2.name,sheet2.nrows,sheet2.ncols
  
        rows = sheet2.row_values(3) # 获取第四行内容
        cols = sheet2.col_values(2) # 获取第三列内容
        print rows
        print cols
  
        print sheet2.cell(1,0).value.encode('utf-8')
        print sheet2.cell_value(1,0).encode('utf-8')
        print sheet2.row(1)[0].value.encode('utf-8')
  
        print sheet2.cell(1,0).ctype
  
